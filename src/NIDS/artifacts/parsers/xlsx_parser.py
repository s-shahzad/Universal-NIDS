from __future__ import annotations

import zipfile
from pathlib import Path
from typing import Any


def parse_xlsx(path: Path, sample_rows: int = 5) -> dict[str, Any]:
    """Extract workbook shape and sample rows from XLSX without macros execution."""
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        return {
            "metadata": {"error": f"openpyxl unavailable: {exc}"},
            "text": "",
            "tags": ["xlsx", "parse_error"],
            "reasons": ["xlsx_parser_unavailable"],
        }

    reasons: list[str] = []

    try:
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        sheet_info: list[dict[str, Any]] = []
        text_snippets: list[str] = []

        for idx, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]
            row_count = int(ws.max_row or 0)
            col_count = int(ws.max_column or 0)
            sheet_info.append(
                {
                    "sheet": sheet_name,
                    "rows": row_count,
                    "cols": col_count,
                }
            )

            if idx == 0:
                for row in ws.iter_rows(min_row=1, max_row=sample_rows, values_only=True):
                    values = ["" if value is None else str(value) for value in row]
                    text_snippets.append(",".join(values))

        wb.close()

        metadata: dict[str, Any] = {"sheets": sheet_info, "sheet_count": len(sheet_info)}

        try:
            with zipfile.ZipFile(path) as archive:
                entries = [item.filename.lower() for item in archive.infolist()]
            has_vba = any("vba" in entry for entry in entries)
            has_external = any("externalLinks".lower() in entry.lower() for entry in entries)
            metadata["has_vba_indicator"] = has_vba
            metadata["has_external_links"] = has_external
            if has_vba:
                reasons.append("xlsx_vba_macro_indicator")
            if has_external:
                reasons.append("xlsx_external_link_indicator")
        except Exception:
            pass

        return {
            "metadata": metadata,
            "text": "\n".join(text_snippets)[:20000],
            "tags": ["xlsx"],
            "reasons": reasons,
        }
    except Exception as exc:
        return {
            "metadata": {"error": f"xlsx_parse_failed: {exc}"},
            "text": "",
            "tags": ["xlsx", "parse_error"],
            "reasons": ["xlsx_parse_failed"],
        }
